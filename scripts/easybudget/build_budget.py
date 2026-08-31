#!/usr/bin/env python3
"""Build the EASY budget: one spec -> budget.xlsx (live formulae) + budget .txt.

Single source of truth for the AP-RCA Western South Sudan support plan budget.

Every rate is a stated PLANNING ASSUMPTION for this landscape, with the note
saying what it assumes and how firm it is. None of them is a quotation and none
is a promise: they are the numbers a lean operation of this shape costs in this
region, to be replaced line by line as real quotes arrive.

Operating model the numbers describe: implementation through local partners and
the South Sudan Wildlife Service, oversight from Chinko HQ, Juba handled by
African Parks South Sudan's country representative. Movement is CAR <-> South
Sudan only; no European travel is budgeted.

Structure and discipline follow the MSF pattern: cost categories, an explicit
unit for every line, quantities derived from a stated field calendar, and
support charged as a rate rather than run as a parallel structure.

Every budget line is a formula: unit cost is a VLOOKUP into Rates, line totals
are products, category subtotals are SUMIFs, support and contingency are
percentage cells. Change a rate once and the whole workbook moves; nothing
downstream is a typed number (AGENTS.md invariant 2).

Usage: python3 scripts/easybudget/build_budget.py
"""
import argparse
import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# RATES: code, label, unit, USD, note (what the rate assumes; firmness)
# --------------------------------------------------------------------------
RATES = [
    # -- people ------------------------------------------------------------
    ("SAL_SCOUT", "Community wildlife scout (ECHO/TANGO), all-in", "person-month", 360,
     "Seasonal community scout on a national contract: basic pay plus statutory "
     "contributions and end-of-season bonus, excluding the food allowance which "
     "is budgeted separately. FIRM to about +/-15%"),
    ("SAL_LEAD", "Scout team leader, all-in", "person-month", 550,
     "One leader per four-person team, roughly half again a member's package. "
     "FIRM to about +/-15%"),
    ("SAL_OFFICER", "National project officer (field liaison, admin)", "person-month", 850,
     "Experienced national officer able to work with county authorities and the "
     "wildlife service unsupervised. INDICATIVE - set against offers received"),
    ("SAL_COORD", "National project coordinator, Wau", "person-month", 1400,
     "Senior national hire carrying a government-facing mandate; the single most "
     "important recruitment in the plan. INDICATIVE - expect to pay to keep"),
    ("SAL_INTL", "Technical adviser from Chinko HQ, all-in (excl. travel)", "person-month", 5500,
     "Charged only for months actually worked on this landscape: salary, "
     "insurance and medevac cover. Not a resident expatriate post. INDICATIVE"),
    ("SSWS_ALLOW", "SSWS staff field allowance, joint patrol/mission", "person-day", 25,
     "Allowance for wildlife-service staff on joint missions. It is an "
     "allowance and never a government salary - the distinction matters to both "
     "the ministry and the donor. FIRM"),
    ("RATION", "Field food allowance", "person-month", 155,
     "Cash food allowance per scout per month in the field, at prevailing "
     "market prices in the western towns. FIRM"),
    ("PERDIEM_F", "Per diem, field mission", "person-day", 15, "Field rate. FIRM"),
    ("PERDIEM_J", "Per diem, Juba/Wau", "person-day", 40, "Town rate. FIRM"),
    ("HEALTH_N", "National staff healthcare/medevac contribution", "person-month", 20,
     "Per capita contribution to treatment, referral and evacuation. FIRM"),
    ("FACILITATOR", "Conservancy-constitution facilitator (national consultant)", "day", 250,
     "National consultant day rate for facilitating a community constitution and "
     "collective-decision rules. INDICATIVE"),
    # -- movement ----------------------------------------------------------
    ("FLT_REG", "Regional return flight, Bangui - Juba (via Addis/Nairobi)", "flight", 1400,
     "There is no direct routing; the trip is two regional legs each way. "
     "Movement is CAR <-> South Sudan only - no European travel is budgeted. "
     "INDICATIVE, volatile with season"),
    ("FLT_DOM", "Domestic flight, one way (Juba - Wau/Tambura)", "flight", 440,
     "Scheduled domestic leg. INDICATIVE, volatile"),
    ("CHARTER", "Caravan charter, one rotation", "flight", 4200,
     "A single Caravan rotation out to the corridor with team and cargo. "
     "INDICATIVE - the cheapest way to reach ground the roads do not"),
    ("FLY_HR", "Light-aircraft flying hour (C182/C206/ULM), wet", "hour", 930,
     "Wet hourly cost for a light aircraft on reconnaissance or survey. A "
     "microlight is roughly half this and a helicopter roughly double. INDICATIVE"),
    ("VEH_HIRE", "4x4 with driver, local hire", "day", 150,
     "Hired, not bought: at this scale a fleet is a liability, not an asset. FIRM"),
    ("MBIKE", "Motorbike, 125cc, delivered and registered", "unit", 4100,
     "Delivered price in South Sudan including transport, registration and "
     "plates. Note this is several times a capital-city street price elsewhere "
     "in the region - do not budget this line from a Bangui or Nairobi quote. FIRM"),
    ("MBIKE_RUN", "Motorbike running (fuel, oil, maintenance)", "vehicle-month", 120,
     "Per bike per active month. FIRM"),
    ("FREIGHT_PCT", "Freight, customs and clearing, applied to imported goods", "share", 0.40,
     "Landed cost of anything bought outside the country runs far above its "
     "invoice: air freight, customs, clearing and inland transport. Budgeting "
     "under 40% on this ground is fiction. FIRM"),
    # -- equipment ---------------------------------------------------------
    ("PHONE", "Rugged smartphone (data collection)", "unit", 250, "Landed. FIRM"),
    ("POWER", "Solar power bank / field charger", "unit", 60, "Landed. FIRM"),
    ("LAPTOP", "Laptop (business class), landed", "unit", 1700,
     "Business-class machine landed in country. FIRM"),
    ("INREACH", "Satellite messenger with mount", "unit", 450,
     "Deliberately a messenger, not a satellite PTT handset: the handsets cost "
     "three times this and the teams need position and a panic button, not "
     "voice. FIRM"),
    ("INREACH_SUB", "Satellite messenger subscription", "unit-month", 35, "FIRM"),
    ("STARLINK_KIT", "Starlink terminal", "unit", 500, "Current hardware price, 2026"),
    ("STARLINK_MO", "Starlink subscription", "unit-month", 40,
     "Current 2026 tariff. Older programmes in the region still carry legacy "
     "business plans at USD 700-900/month - do not copy those figures across"),
    ("AIRTIME", "Mobile airtime", "person-month", 15, "Per staff member. FIRM"),
    ("UNIFORM", "Scout uniform and personal kit", "person-year", 250,
     "Two sets, boots, pack and badges, excluding freight. FIRM"),
    ("FIELDKIT", "Team field kit (tent, tarp, cooking, torches, pangas)", "team", 1400,
     "Per four-person team, excluding freight. FIRM"),
    ("CAMTRAP", "Camera trap", "unit", 450, "Landed, per unit. FIRM"),
    ("GEN", "Generator, 5 kVA", "unit", 2500, "Installed. FIRM"),
    ("SOLAR", "Office solar and battery backup", "set", 4000, "Installed. INDICATIVE"),
    # -- premises and running ----------------------------------------------
    ("OFFICE_WAU", "Office/compound rent, Wau, shared with the national partner", "month", 800,
     "A shared room and yard in the partner's existing premises. No Juba office "
     "is budgeted: African Parks South Sudan hosts the Juba desk. FIRM"),
    ("OFFICE_RUN", "Office running (fuel, power, water, cleaning, consumables)", "month", 350,
     "FIRM"),
    ("ACC_JUBA", "Accommodation, Juba", "night", 100, "FIRM"),
    ("ACC_FIELD", "Accommodation, field/town", "night", 35, "FIRM"),
    # -- programme ----------------------------------------------------------
    ("TRAIN_PD", "Training, per participant-day (venue, board, allowance, materials)",
     "person-day", 30,
     "All-in per participant per training day: board and lodging, the trainee "
     "allowance, hall hire and materials. FIRM"),
    ("MEETING", "Community/stakeholder meeting (refreshments, transport, hall)", "meeting", 450,
     "One community or leadership meeting including participants' transport. FIRM"),
    ("BOREHOLE", "Borehole, drilled and equipped", "unit", 16900,
     "Drilled, cased, hand-pump equipped, with a water-committee handover. "
     "INDICATIVE - depth and haulage distance drive it"),
    ("VET_CAMP", "Veterinary campaign, one dry season (drugs, vaccination, handling)",
     "campaign", 14000,
     "One dry-season campaign along the corridor, delivered by the veterinary "
     "partner. INDICATIVE"),
    ("SURVEY_TECH", "Aerial survey technical package (design, crew, analysis)", "survey", 25000,
     "Design, observers, data handling and the written result - flying hours "
     "are a separate line. Assumes the aircraft and protocol are BORROWED from "
     "an operator that already has them, not procured. INDICATIVE"),
    ("LEGAL", "Legal retainer, in-country counsel", "month", 500,
     "Monthly retainer, not a per-matter fee. FIRM"),
    ("REG_NGO", "NGO/CBO registration, renewals, no-objection letters", "year", 3500,
     "Annual cycle of registration, renewal and the letters that unlock "
     "operations. FIRM"),
    ("PERMIT", "Work and residence permit, per expatriate", "person-year", 1000,
     "NGO Act 2016 s.18 regime. FIRM"),
    ("AUDIT", "Annual audit and statutory accounts", "year", 12000,
     "Independent audit, required at renewal by NGO Act 2016 s.13. FIRM"),
    ("GRANT_HARD", "Partner grant: national NGO, Wau/Raga base and mobilisation", "year", 60000,
     "The implementing body on the ground: premises, mobilisers, county-level "
     "registration and community logistics in both anchor towns. Buying into an "
     "organisation that is already registered and already present is cheaper "
     "and faster than building one. INDICATIVE - set by negotiation"),
    ("GRANT_VET", "Partner grant: veterinary delivery on the corridor", "year", 35000,
     "The veterinary half of the price the herders named. INDICATIVE"),
    ("GRANT_LOCAL", "Partner grant: local CBO at a single site", "year", 15000,
     "A county-registered local body on the south-east approach. INDICATIVE"),
    ("SSWS_SUPP", "SSWS institutional support (posts, fuel, radios, office kit)", "year", 20000,
     "Equipping the two existing but unequipped posts and the state office: "
     "radios and solar, fuel, a computer, connectivity. This is what makes SSWS "
     "the visible lead rather than a logo. INDICATIVE"),
]

# --------------------------------------------------------------------------
# BUDGET LINES: category, action, item, basis, rate code, (Y1, Y2, Y3)
# --------------------------------------------------------------------------
CATS = [
    "1 International staff",
    "2 National staff",
    "3 Local partners and SSWS",
    "4 Field activities",
    "5 Equipment and supplies",
    "6 Transport",
    "7 Travel and accommodation",
    "8 Premises and utilities",
    "9 Communications and IT",
    "10 Training",
    "11 Legal, registration and audit",
]

GOODS_CATS = ("5 Equipment and supplies", "6 Transport")

LINES = [
    # ---- 1 International staff -------------------------------------------
    ("1 International staff", "A1",
     "Technical adviser, based Chinko HQ (oversight visits plus remote)",
     "3 months/yr Y1-Y2, 2 in Y3. Oversight from Chinko, not a resident "
     "expatriate structure in South Sudan", "SAL_INTL", (3, 3, 2), 0.67, "HQ"),
    ("1 International staff", "A1", "Adviser work and residence permits",
     "NGO Act 2016 s.18", "PERMIT", (1, 1, 1), 1.0, "HQ"),
    # ---- 2 National staff --------------------------------------------------
    ("2 National staff", "A1", "Project coordinator, Wau (national)",
     "12 months/yr from month 1 - the permanent face of the project on the ground",
     "SAL_COORD", (12, 12, 12), 0.5, "FP"),
    ("2 National staff", "A1", "Field liaison officer, Raga / Deim Zubeir (national)",
     "One officer Y1, two from Y2", "SAL_OFFICER", (12, 24, 24), 0.5, "FP"),
    ("2 National staff", "A1", "Finance/admin officer, 0.5 FTE hosted by AP South Sudan, Juba",
     "Half-time. Juba admin, banking and approvals run through the AP South "
     "Sudan country representative's office rather than a second back office",
     "SAL_OFFICER", (6, 6, 6), 0.5, "FP"),
    ("2 National staff", "A4", "Community wildlife scouts (ECHO/TANGO members)",
     "Seasonal contracts: Y1 3 teams x 4 x 5 months (Nov-Mar); Y2 4 x 4 x 7; "
     "Y3 4 x 4 x 9. The audience is seasonal (assessment section 5)",
     "SAL_SCOUT", (60, 112, 144), 1.0, "TEAM"),
    ("2 National staff", "A4", "Scout team leaders", "One per team, same months",
     "SAL_LEAD", (15, 28, 36), 1.0, "TEAM"),
    ("2 National staff", "A4", "Scout and leader food allowance",
     "Same person-months as scouts plus leaders", "RATION", (75, 140, 180), 1.0, "TEAM"),
    ("2 National staff", "A1", "National staff healthcare contribution",
     "All national person-months", "HEALTH_N", (105, 182, 222), 0.6, "TEAM"),
    # ---- 3 Local partners and SSWS ----------------------------------------
    ("3 Local partners and SSWS", "A4",
     "Partner grant: national NGO, Wau/Raga base and mobilisation",
     "The implementing body on the ground. Registered, present in both anchor "
     "towns, and already inside the NGO Act's staffing rule", "GRANT_HARD", (1, 1, 1), 0.5, "FP"),
    ("3 Local partners and SSWS", "A6",
     "Partner grant: veterinary delivery on the corridor",
     "From Y2, once the corridor talks have a counterpart and a route",
     "GRANT_VET", (0, 1, 1), 0.5, None),
    ("3 Local partners and SSWS", "A9",
     "Partner grant: local CBO, Nagero / south-east approach", "From Y2",
     "GRANT_LOCAL", (0, 1, 1), 0.5, "FP"),
    ("3 Local partners and SSWS", "A1",
     "SSWS institutional support (Deim Zubeir and Boro-Medina posts, radios, fuel)",
     "Visible SSWS ownership is the premise of the whole plan (action 1); the "
     "posts exist and are unequipped", "SSWS_SUPP", (1, 1, 1), 0.7, "FP"),
    ("3 Local partners and SSWS", "A1", "SSWS field allowances, joint patrols and missions",
     "Y1 6 staff x 30 days; Y2 8 x 45; Y3 8 x 60. Teams joint or SSWS-embedded, "
     "never parallel", "SSWS_ALLOW", (180, 360, 480), 1.0, "TEAM"),
    # ---- 4 Field activities ------------------------------------------------
    ("4 Field activities", "A7",
     "Rim reconnaissance flights (Busseri headwaters, Nahr al Jur)",
     "2 flights/yr x 6 hours in the Dec-Feb window. Mines here are dark to "
     "satellites; this plus the ground question is the whole monitoring system",
     "FLY_HR", (12, 12, 12), 1.0, None),
    ("4 Field activities", "A7",
     "Aerial wildlife survey, Numatina-Boro: technical package",
     "Y2 only. First survey since 2007. Capability borrowed from African Parks "
     "or FFI/Bucknell, not procured", "SURVEY_TECH", (0, 1, 0), 0.0, None),
    ("4 Field activities", "A7", "Aerial survey flying hours",
     "Y2: 40 hours over ~14,500 km2 at survey altitude", "FLY_HR", (0, 40, 0), 0.0, None),
    ("4 Field activities", "A7",
     "Ground field-check missions (rim cells; the three uncertain sites)",
     "Y1 4 missions x 12 vehicle-days; Y2-Y3 3 x 12", "VEH_HIRE", (48, 36, 36), 1.0, None),
    ("4 Field activities", "A6",
     "Corridor and conservancy meetings with herder leadership and communities",
     "Y1 12 meetings, Dec-Feb only - talks outside that window happen with "
     "nobody; Y2 20; Y3 24", "MEETING", (12, 20, 24), 1.0, "FP"),
    ("4 Field activities", "A6", "Water points on the corridor (boreholes)",
     "Y2 one, Y3 two. The herders named the price: designated ground, water, "
     "veterinary and medical support - delivered, not promised",
     "BOREHOLE", (0, 1, 2), 0.0, None),
    ("4 Field activities", "A6", "Veterinary campaign, dry season",
     "Y2 and Y3, delivered through the veterinary partner", "VET_CAMP", (0, 1, 1), 0.0, None),
    ("4 Field activities", "A4", "Conservancy constitution facilitation",
     "Y1 30 days (two conservancies scoped); Y2 60 (applications filed); Y3 40. "
     "The constitution is the slow part - a facilitator, not a lawyer's afternoon",
     "FACILITATOR", (30, 60, 40), 0.5, None),
    ("4 Field activities", "A10", "Camera traps for community biomonitoring",
     "Y2 30 units on the academic partner's existing protocol; Y3 10 replacements",
     "CAMTRAP", (0, 30, 10), 0.0, "TEAM"),
    # ---- 5 Equipment and supplies -----------------------------------------
    ("5 Equipment and supplies", "A4", "Scout uniforms and personal kit",
     "Per scout and leader per year", "UNIFORM", (15, 20, 20), 1.0, "TEAM"),
    ("5 Equipment and supplies", "A4", "Team field kit",
     "Y1 3 teams; Y2 one new team plus a replacement; Y3 replacements",
     "FIELDKIT", (3, 2, 2), 1.0, "TEAM"),
    ("5 Equipment and supplies", "A10", "Rugged smartphones for the ground-truth checklist",
     "One per scout pair plus coordinator; top-ups in Y2-Y3", "PHONE", (10, 8, 6), 1.0, "TEAM"),
    ("5 Equipment and supplies", "A10", "Solar power banks", "One with each phone",
     "POWER", (10, 8, 6), 1.0, "TEAM"),
    ("5 Equipment and supplies", "A10", "Laptops (coordinator, liaison officer, partner)",
     "Y1 3; Y3 2 replacements", "LAPTOP", (3, 0, 2), 1.0, "FP"),
    ("5 Equipment and supplies", "A1", "Generator, Wau base", "Y1 only", "GEN", (1, 0, 0), 1.0, "FP"),
    ("5 Equipment and supplies", "A1", "Solar and battery backup, Wau base", "Y1 only",
     "SOLAR", (1, 0, 0), 1.0, "FP"),
    # ---- 6 Transport --------------------------------------------------------
    ("6 Transport", "A4", "Motorbikes for team movement",
     "Y1 3, one per team; Y2 one for the new team", "MBIKE", (3, 1, 0), 1.0, "TEAM"),
    ("6 Transport", "A4", "Motorbike running costs",
     "Vehicle-months: 3 bikes x 5 months Y1; 4 x 7 Y2; 4 x 9 Y3",
     "MBIKE_RUN", (15, 28, 36), 1.0, "TEAM"),
    ("6 Transport", "A4", "Charter rotations (teams and equipment to the corridor)",
     "Y1 2; Y2 4 (survey and borehole logistics); Y3 3", "CHARTER", (2, 4, 3), 1.0, "TEAM"),
    # ---- 7 Travel and accommodation ---------------------------------------
    ("7 Travel and accommodation", "A1",
     "Regional flights, Bangui - Juba (adviser and Chinko oversight)",
     "3 rotations/yr Y1-Y2, 2 in Y3. CAR <-> South Sudan only", "FLT_REG", (3, 3, 2), 0.67, "HQ"),
    ("7 Travel and accommodation", "A1", "Domestic flights, Juba - Wau/Tambura",
     "Y1 24 legs; Y2-Y3 30", "FLT_DOM", (24, 30, 30), 0.6, "FP"),
    ("7 Travel and accommodation", "A2",
     "Juba accommodation (paper track: regulations, gazette request, RRC)",
     "Y1 60 nights - ten weeks of paper time before S1 opens; Y2-Y3 40",
     "ACC_JUBA", (60, 40, 40), 0.83, "FP"),
    ("7 Travel and accommodation", "A2", "Juba per diems", "Same days as above",
     "PERDIEM_J", (60, 40, 40), 0.83, "FP"),
    ("7 Travel and accommodation", "A7", "Field accommodation, missions",
     "Y1 120 person-nights; Y2-Y3 150", "ACC_FIELD", (120, 150, 150), 1.0, None),
    ("7 Travel and accommodation", "A7", "Field per diems", "Same days as above",
     "PERDIEM_F", (120, 150, 150), 1.0, None),
    # ---- 8 Premises and utilities ------------------------------------------
    ("8 Premises and utilities", "A1",
     "Wau office/compound, shared with the national partner",
     "12 months/yr. No Juba office: AP South Sudan hosts the desk",
     "OFFICE_WAU", (12, 12, 12), 0.5, "FP"),
    ("8 Premises and utilities", "A1", "Office running costs", "12 months/yr",
     "OFFICE_RUN", (12, 12, 12), 0.5, "FP"),
    # ---- 9 Communications and IT -------------------------------------------
    ("9 Communications and IT", "A1", "Starlink terminals (Wau base, forward post)",
     "Y1 2; Y2 one more", "STARLINK_KIT", (2, 1, 0), 1.0, "FP"),
    ("9 Communications and IT", "A1", "Starlink subscriptions",
     "Unit-months: 2 x 12 in Y1, 3 x 12 in Y2-Y3", "STARLINK_MO", (24, 36, 36), 0.5, "FP"),
    ("9 Communications and IT", "A10",
     "Satellite messengers (team safety and track logging)",
     "One per team plus coordinator", "INREACH", (4, 2, 1), 1.0, "TEAM"),
    ("9 Communications and IT", "A10", "Satellite messenger subscriptions",
     "Unit-months, seasonal use", "INREACH_SUB", (20, 35, 45), 1.0, "TEAM"),
    ("9 Communications and IT", "A1", "Mobile airtime", "All staff person-months",
     "AIRTIME", (105, 182, 222), 0.6, "FP"),
    # ---- 10 Training --------------------------------------------------------
    ("10 Training", "A4", "Scout induction and community-engagement training",
     "Y1 15 participants x 14 days; Y2 20 x 14; Y3 20 x 7 refresher. Community "
     "engagement is explicit: the wildlife service descends from an armed force",
     "TRAIN_PD", (210, 280, 140), 1.0, "TEAM"),
    ("10 Training", "A10",
     "Ground-truth checklist and data training (fire, gold, herd routes, river names)",
     "Y1 20 x 3 days; Y2-Y3 25 x 3", "TRAIN_PD", (60, 75, 75), 1.0, "TEAM"),
    ("10 Training", "A8",
     "Land-use planning workshops with the boom towns north-east of the park",
     "Y2 and Y3: 4 workshops x 25 participants x 2 days, while their cropland "
     "is still ~1.5%", "TRAIN_PD", (0, 200, 200), 0.0, None),
    # ---- 11 Legal, registration, audit --------------------------------------
    ("11 Legal, registration and audit", "A2",
     "In-country counsel: regulations text, s.24 gazette request, boundary description",
     "Retainer, 12 months/yr. Actions 2 and 3 are the cheapest and most durable "
     "acts in the plan", "LEGAL", (12, 12, 12), 0.5, None),
    ("11 Legal, registration and audit", "A4",
     "NGO/CBO registration, RRC renewal, no-objection letters",
     "Annual. Two clocks, both slow - start them in parallel", "REG_NGO", (1, 1, 1), 1.0, None),
    ("11 Legal, registration and audit", "A1", "Annual audit and statutory accounts",
     "NGO Act 2016 s.13 requires audited accounts at renewal", "AUDIT", (1, 1, 1), 0.0, None),
]

ACTIONS = {
    "A1": "1  Secure SSWS backing first",
    "A2": "2  Get into the regulations being drafted",
    "A3": "3  Ask for the s.24 gazette closure order",
    "A4": "4  Register under the 2026 Act early, rim first",
    "A5": "5  Treat the road answer as final",
    "A6": "6  Keep the corridor axis, budget its price",
    "A7": "7  Field-check and fly the rim",
    "A8": "8  Land-use planning with the boom towns",
    "A9": "9  Fix the site list",
    "A10": "10 Ground-truth checklist for the teams",
    "A11": "11 Attach to the funding architecture",
}

ASSUMPTIONS = [
    ("FREIGHT_PCT", "Freight, customs and clearing on equipment and transport", 0.40,
     "Applied to categories 5 and 6. Anything imported arrives at well above its "
     "invoice: air freight, customs, clearing agent and inland transport. On "
     "light bulky goods - uniforms, tents, camera traps - the delivery cost "
     "approaches the goods cost. Budgeting under 40% here is fiction"),
    ("SUPPORT_PCT", "Chinko HQ oversight and AP South Sudan Juba desk", 0.08,
     "Shared services actually consumed: finance and payroll systems, "
     "procurement, Juba representation and approvals, HR, safety. Charged as a "
     "rate against a service the project genuinely uses rather than run as a "
     "second structure - which is the whole point of hanging this off two "
     "existing operations"),
    ("CONTING_PCT", "Contingency", 0.07,
     "One field window a year (15 Dec - 15 Feb) in a low-frequency, "
     "high-amplitude state: the risk being priced is a lost season, plus SSP/USD "
     "movement on locally settled costs"),
    ("BANK_PCT", "Bank charges and FX cost", 0.015,
     "Transfer fees and the spread on moving USD into a country that settles "
     "much of this budget in local currency"),
]

# --------------------------------------------------------------------------
# workbook
# --------------------------------------------------------------------------
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(bottom=THIN)
H_FILL = PatternFill("solid", fgColor="1F3864")
C_FILL = PatternFill("solid", fgColor="D9E2F3")
T_FILL = PatternFill("solid", fgColor="FFE699")
MONEY = '#,##0'


def style_header(ws, row, ncol):
    for i in range(1, ncol + 1):
        c = ws.cell(row=row, column=i)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = H_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def build_xlsx(path):
    wb = openpyxl.Workbook()

    # ---------------- Rates ------------------------------------------------
    ws = wb.active
    ws.title = "Rates"
    ws["A1"] = "UNIT COST CATALOGUE - every rate below is used by formula in the Budget sheet"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("Every rate below is a stated planning assumption for this landscape, "
                "marked FIRM (expect to pay about this) or INDICATIVE (replace with a "
                "quote before signing). None is a quotation and none is a promise.")
    ws["A2"].font = Font(italic=True, size=9)
    hdr = ["Code", "Item", "Unit", "Unit cost USD", "What the rate assumes / how firm it is"]
    ws.append([])
    ws.append(hdr)
    style_header(ws, 4, len(hdr))
    for code, label, unit, usd, src in RATES:
        ws.append([code, label, unit, usd, src])
        r = ws.max_row
        ws.cell(row=r, column=4).number_format = '#,##0.000' if usd < 1 else MONEY
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        for i in range(1, 6):
            ws.cell(row=r, column=i).border = BORDER
    rate_first, rate_last = 5, ws.max_row
    for col, w in zip("ABCDE", (14, 46, 16, 14, 78)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    RANGE = f"Rates!$A${rate_first}:$D${rate_last}"

    # ---------------- Assumptions -----------------------------------------
    wa = wb.create_sheet("Assumptions")
    wa["A1"] = "RATES APPLIED TO THE WHOLE BUDGET"
    wa["A1"].font = Font(bold=True, size=12)
    wa.append([])
    wa.append(["Code", "What it covers", "Rate", "Basis"])
    style_header(wa, 3, 4)
    keycell = {}
    for code, label, val, src in ASSUMPTIONS:
        wa.append([code, label, val, src])
        r = wa.max_row
        wa.cell(row=r, column=3).number_format = '0.0%'
        wa.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        wa.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        keycell[code] = f"Assumptions!$C${r}"
    for col, w in zip("ABCD", (16, 52, 10, 86)):
        wa.column_dimensions[col].width = w

    # ---------------- Budget ----------------------------------------------
    wb_ = wb.create_sheet("Budget")
    wb_["A1"] = ("AP-RCA WESTERN SOUTH SUDAN - ASSESSMENT TO ACTION: THREE-YEAR BUDGET")
    wb_["A1"].font = Font(bold=True, size=13)
    wb_["A2"] = ("Implementation through local partners and SSWS; oversight from Chinko HQ; "
                 "Juba handled by African Parks South Sudan. Unit costs are formulae into "
                 "the Rates sheet - change a rate there, not here.")
    wb_["A2"].font = Font(italic=True, size=9)
    hdr = ["Category", "Plan action", "Line item", "Basis / quantity logic", "Rate code",
           "Unit", "Unit cost USD", "Qty Y1", "Total Y1", "Qty Y2", "Total Y2",
           "Qty Y3", "Total Y3", "Total 3 yr", "Share in first 6 months",
           "Of which first 6 months", "Delivery unit"]
    wb_.append([])
    wb_.append(hdr)
    style_header(wb_, 4, len(hdr))
    first = 5
    for cat, act, item, basis, code, qty, h1, drv in LINES:
        wb_.append([cat, ACTIONS[act], item, basis, code, "", "", qty[0], "",
                    qty[1], "", qty[2], "", "", h1, "", drv or ""])
        r = wb_.max_row
        wb_.cell(row=r, column=6).value = f'=VLOOKUP($E{r},Rates!$A${rate_first}:$E${rate_last},3,FALSE)'
        wb_.cell(row=r, column=7).value = f'=VLOOKUP($E{r},{RANGE},4,FALSE)'
        wb_.cell(row=r, column=9).value = f'=H{r}*$G{r}'
        wb_.cell(row=r, column=11).value = f'=J{r}*$G{r}'
        wb_.cell(row=r, column=13).value = f'=L{r}*$G{r}'
        wb_.cell(row=r, column=14).value = f'=I{r}+K{r}+M{r}'
        wb_.cell(row=r, column=15).number_format = '0%'
        wb_.cell(row=r, column=16).value = f'=I{r}*O{r}'
        for c in (7, 9, 11, 13, 14, 16):
            wb_.cell(row=r, column=c).number_format = MONEY
        for c in (3, 4):
            wb_.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, 18):
            wb_.cell(row=r, column=c).border = BORDER
    last = wb_.max_row

    def money_row(label, formulas, bold=True, fill=None, fmt=MONEY):
        wb_.append([label])
        r = wb_.max_row
        for col, f in formulas.items():
            cell = wb_.cell(row=r, column=col)
            cell.value = f
            cell.number_format = fmt
        for c in range(1, 18):
            cc = wb_.cell(row=r, column=c)
            if bold:
                cc.font = Font(bold=True)
            if fill:
                cc.fill = fill
        return r

    wb_.append([])
    r_direct = money_row("DIRECT COSTS (categories 1-11)", {
        9: f"=SUM(I{first}:I{last})", 11: f"=SUM(K{first}:K{last})",
        13: f"=SUM(M{first}:M{last})", 14: f"=SUM(N{first}:N{last})",
        16: f"=SUM(P{first}:P{last})"}, fill=C_FILL)
    goods_crit = "+".join(
        f'SUMIF($A${first}:$A${last},"{g}",{{c}}${first}:{{c}}${last})' for g in GOODS_CATS)
    r_freight = money_row("Freight, customs and clearing on categories 5-6", {
        9: "=(" + goods_crit.format(c="I") + f")*{keycell['FREIGHT_PCT']}",
        11: "=(" + goods_crit.format(c="K") + f")*{keycell['FREIGHT_PCT']}",
        13: "=(" + goods_crit.format(c="M") + f")*{keycell['FREIGHT_PCT']}",
    }, bold=False)
    wb_.cell(row=r_freight, column=14).value = f"=I{r_freight}+K{r_freight}+M{r_freight}"
    wb_.cell(row=r_freight, column=14).number_format = MONEY
    r_bank = money_row("Bank charges and FX cost", {
        9: f"=(I{r_direct}+I{r_freight})*{keycell['BANK_PCT']}",
        11: f"=(K{r_direct}+K{r_freight})*{keycell['BANK_PCT']}",
        13: f"=(M{r_direct}+M{r_freight})*{keycell['BANK_PCT']}"}, bold=False)
    for col in (9, 11, 13):
        wb_.cell(row=r_bank, column=col).number_format = MONEY
    wb_.cell(row=r_bank, column=14).value = f"=I{r_bank}+K{r_bank}+M{r_bank}"
    wb_.cell(row=r_bank, column=14).number_format = MONEY
    r_sub = money_row("SUBTOTAL, FIELD COST", {
        9: f"=I{r_direct}+I{r_freight}+I{r_bank}",
        11: f"=K{r_direct}+K{r_freight}+K{r_bank}",
        13: f"=M{r_direct}+M{r_freight}+M{r_bank}",
        14: f"=N{r_direct}+N{r_freight}+N{r_bank}"}, fill=C_FILL)
    r_supp = money_row("Chinko HQ oversight + AP South Sudan Juba desk", {
        9: f"=I{r_sub}*{keycell['SUPPORT_PCT']}",
        11: f"=K{r_sub}*{keycell['SUPPORT_PCT']}",
        13: f"=M{r_sub}*{keycell['SUPPORT_PCT']}",
        14: f"=N{r_sub}*{keycell['SUPPORT_PCT']}"}, bold=False)
    r_cont = money_row("Contingency", {
        9: f"=(I{r_sub}+I{r_supp})*{keycell['CONTING_PCT']}",
        11: f"=(K{r_sub}+K{r_supp})*{keycell['CONTING_PCT']}",
        13: f"=(M{r_sub}+M{r_supp})*{keycell['CONTING_PCT']}",
        14: f"=(N{r_sub}+N{r_supp})*{keycell['CONTING_PCT']}"}, bold=False)
    r_tot = money_row("TOTAL REQUESTED, USD", {
        9: f"=I{r_sub}+I{r_supp}+I{r_cont}",
        11: f"=K{r_sub}+K{r_supp}+K{r_cont}",
        13: f"=M{r_sub}+M{r_supp}+M{r_cont}",
        14: f"=N{r_sub}+N{r_supp}+N{r_cont}"}, fill=T_FILL)

    goods_h1 = "+".join(
        f'SUMPRODUCT(($A${first}:$A${last}="{g}")*$I${first}:$I${last}'
        f'*$O${first}:$O${last})' for g in GOODS_CATS)
    r_h1_direct = money_row("FIRST SIX MONTHS (Oct 2026 - Mar 2027), direct", {
        9: f"=SUM(P{first}:P{last})"}, bold=False)
    r_h1_freight = money_row("  ...its freight, customs and clearing", {
        9: f"=({goods_h1})*{keycell['FREIGHT_PCT']}"}, bold=False)
    r_h1_bank = money_row("  ...its bank charges and FX", {
        9: f"=(I{r_h1_direct}+I{r_h1_freight})*{keycell['BANK_PCT']}"}, bold=False)
    r_h1 = money_row("FIRST SIX MONTHS, LOADED TOTAL", {
        9: f"=(I{r_h1_direct}+I{r_h1_freight}+I{r_h1_bank})"
           f"*(1+{keycell['SUPPORT_PCT']})*(1+{keycell['CONTING_PCT']})"},
        fill=C_FILL)

    for col, w in zip("ABCDEFGHIJKLMNOPQ",
                      (26, 34, 52, 60, 13, 14, 12, 8, 12, 8, 12, 8, 12, 13, 10, 14, 12)):
        wb_.column_dimensions[col].width = w
    wb_.freeze_panes = "C5"

    # ---------------- Summary by category ---------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "SUMMARY BY COST CATEGORY (all figures are formulae over the Budget sheet)"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.append([])
    ws2.append(["Category", "Y1", "Y2", "Y3", "Total 3 yr", "% of direct"])
    style_header(ws2, 3, 6)
    for cat in CATS:
        ws2.append([cat])
        r = ws2.max_row
        for i, col in enumerate("IKM", start=2):
            ws2.cell(row=r, column=i).value = (
                f'=SUMIF(Budget!$A${first}:$A${last},$A{r},Budget!${col}${first}:${col}${last})')
            ws2.cell(row=r, column=i).number_format = MONEY
        ws2.cell(row=r, column=5).value = f"=SUM(B{r}:D{r})"
        ws2.cell(row=r, column=5).number_format = MONEY
    cat_first, cat_last = 4, ws2.max_row
    ws2.append(["DIRECT TOTAL"])
    rt = ws2.max_row
    for i in range(2, 6):
        L_ = get_column_letter(i)
        ws2.cell(row=rt, column=i).value = f"=SUM({L_}{cat_first}:{L_}{cat_last})"
        ws2.cell(row=rt, column=i).number_format = MONEY
        ws2.cell(row=rt, column=i).font = Font(bold=True)
    ws2.cell(row=rt, column=1).font = Font(bold=True)
    for r in range(cat_first, cat_last + 1):
        ws2.cell(row=r, column=6).value = f"=IF($E${rt}=0,\"\",E{r}/$E${rt})"
        ws2.cell(row=r, column=6).number_format = '0.0%'
    ws2.append([])
    for label, src in (("Freight, customs, clearing", r_freight),
                       ("Bank charges and FX", r_bank),
                       ("Chinko HQ + AP South Sudan support", r_supp),
                       ("Contingency", r_cont)):
        ws2.append([label])
        r = ws2.max_row
        for i, col in enumerate("IKMN", start=2):
            ws2.cell(row=r, column=i).value = f"=Budget!${col}${src}"
            ws2.cell(row=r, column=i).number_format = MONEY
    ws2.append(["TOTAL REQUESTED, USD"])
    r = ws2.max_row
    for i, col in enumerate("IKMN", start=2):
        c = ws2.cell(row=r, column=i)
        c.value = f"=Budget!${col}${r_tot}"
        c.number_format = MONEY
        c.font = Font(bold=True)
        c.fill = T_FILL
    ws2.cell(row=r, column=1).font = Font(bold=True)
    for col, w in zip("ABCDEF", (40, 14, 14, 14, 16, 12)):
        ws2.column_dimensions[col].width = w

    # ---------------- Summary by plan action ------------------------------
    ws3 = wb.create_sheet("By action")
    ws3["A1"] = "COST BY PLAN ACTION (the eleven actions of the assessment)"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3["A2"] = ("Actions 3, 5 and 11 carry no direct cost: the gazette request, the road "
                 "finding and the funding conversations are signatures and phone calls, "
                 "paid for inside staff time.")
    ws3["A2"].font = Font(italic=True, size=9)
    ws3.append([])
    ws3.append(["Plan action", "Y1", "Y2", "Y3", "Total 3 yr"])
    style_header(ws3, 4, 5)
    for key in sorted(ACTIONS, key=lambda k: int(k[1:])):
        ws3.append([ACTIONS[key]])
        r = ws3.max_row
        for i, col in enumerate("IKM", start=2):
            ws3.cell(row=r, column=i).value = (
                f'=SUMIF(Budget!$B${first}:$B${last},$A{r},Budget!${col}${first}:${col}${last})')
            ws3.cell(row=r, column=i).number_format = MONEY
        ws3.cell(row=r, column=5).value = f"=SUM(B{r}:D{r})"
        ws3.cell(row=r, column=5).number_format = MONEY
    for col, w in zip("ABCDE", (44, 14, 14, 14, 16)):
        ws3.column_dimensions[col].width = w

    # ---------------- Phasing on the plan's own clock ----------------------
    ws4 = wb.create_sheet("Phasing")
    ws4["A1"] = "PHASING ON THE PLAN'S OWN CLOCK: 6 MONTHS, 1 YEAR, 2 YEARS, 3 YEARS AND ON"
    ws4["A1"].font = Font(bold=True, size=12)
    ws4["A2"] = ("Two clocks set every figure: the FIELD clock has one window a year "
                 "(15 Dec - 15 Feb); the PAPER clock runs all year from Juba. The first "
                 "six months (Oct 2026 - Mar 2027) carry ten weeks of paper time plus "
                 "season one. Column O of the Budget sheet holds the share of year one "
                 "that falls in those six months, line by line.")
    ws4["A2"].alignment = Alignment(wrap_text=True)
    ws4.row_dimensions[2].height = 45
    ws4.append([])
    ws4.append(["Period", "What it buys", "Direct", "Loaded total", "Cumulative"])
    style_header(ws4, 4, 5)
    h1_direct = f"=Budget!$I${r_h1_direct}"
    loadf = f"(Budget!$I${r_tot}/Budget!$I${r_direct})"
    periods = [
        ("6 MONTHS  Oct 2026 - Mar 2027",
         "Get legal, get seen, check the rim. Paper filed in Juba; season one "
         "deployed on SSWS backing.", f"=Budget!$I${r_h1_direct}",
         f"=Budget!$I${r_h1}"),
        ("1 YEAR  Oct 2026 - Sep 2027",
         "Turn the season into instruments: conservancy applications, a costed "
         "corridor package, the survey designed and funded.",
         f"=Budget!$I${r_direct}", f"=Budget!$I${r_tot}"),
        ("2 YEARS  Oct 2027 - Sep 2028",
         "First real protection: a conservancy authorized, scouts patrolling "
         "jointly, the survey flown, the first fire number, water on the corridor.",
         f"=Budget!$K${r_direct}", f"=Budget!$K${r_tot}"),
        ("3 YEARS AND ON  Oct 2028 onward",
         "Declare, then hold it: the park package, the conservancy ring, the "
         "corridor operating and measured each dry season.",
         f"=Budget!$M${r_direct}", f"=Budget!$M${r_tot}"),
    ]
    for label, what, dform, tform in periods:
        ws4.append([label, what, dform, tform])
        r = ws4.max_row
        for c in (3, 4, 5):
            ws4.cell(row=r, column=c).number_format = MONEY
        ws4.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    p_first = 5
    p_last = ws4.max_row
    # cumulative: the 6-month row is inside year one, so it does not add
    ws4.cell(row=p_first, column=5).value = f"=D{p_first}"
    ws4.cell(row=p_first + 1, column=5).value = f"=D{p_first+1}"
    ws4.cell(row=p_first + 2, column=5).value = f"=E{p_first+1}+D{p_first+2}"
    ws4.cell(row=p_first + 3, column=5).value = f"=E{p_first+2}+D{p_first+3}"
    for r in range(p_first, p_last + 1):
        ws4.cell(row=r, column=5).number_format = MONEY
    ws4.append([])
    ws4.append(["NOTE", "The 6-month row is a slice of year one, not an addition to it: "
                "the cumulative column therefore starts again at the 1-year row."])
    ws4.cell(row=ws4.max_row, column=2).alignment = Alignment(wrap_text=True)
    for col, w in zip("ABCDE", (34, 62, 14, 15, 15)):
        ws4.column_dimensions[col].width = w

    # ---------------- Loaded cost of a delivery unit -----------------------
    ws5 = wb.create_sheet("Loaded cost")
    ws5["A1"] = "LOADED COST OF EACH DELIVERY UNIT"
    ws5["A1"].font = Font(bold=True, size=12)
    ws5["A2"] = ("Direct cost is what the line items say. LOADED cost adds each unit's "
                 "share of freight, bank charges, Chinko HQ and Juba support, and "
                 "contingency - so these are the numbers to quote when somebody asks "
                 "what a team or a focal point actually costs to run.")
    ws5["A2"].alignment = Alignment(wrap_text=True)
    ws5.row_dimensions[2].height = 45
    ws5.append([])
    ws5.append(["Loading factor (total requested / direct cost, year 1)", "",
                f"=Budget!$I${r_tot}/Budget!$I${r_direct}"])
    ws5.cell(row=ws5.max_row, column=3).number_format = '0.000'
    lf = f"$C${ws5.max_row}"
    ws5.append([])
    ws5.append(["Delivery unit", "Basis", "Direct Y1", "Direct Y2", "Direct Y3",
                "Loaded Y1", "Loaded Y2", "Loaded Y3", "Loaded 3 yr",
                "Loaded per unit per year"])
    style_header(ws5, ws5.max_row, 10)
    units = [
        ("TEAM", "ECHO/TANGO scout teams (all teams)",
         "Scouts, leaders, food allowance, healthcare, uniforms, field kit, "
         "phones, power banks, camera traps, motorbikes and their running, "
         "charters, SSWS joint-mission allowances, induction and checklist training",
         (3, 4, 4), "team"),
        ("FP", "Focal points and the Wau/Juba backbone",
         "Coordinator, liaison officers, half-time finance officer, national and "
         "local partner grants, SSWS institutional support, Wau office and its "
         "running, Starlink, airtime, laptops, generator and solar, domestic "
         "flights, Juba nights and per diems, community meetings",
         (4, 4, 4), "focal point"),
        ("HQ", "Chinko HQ technical oversight",
         "Adviser months actually worked on this landscape, permits and the "
         "regional flights that carry them", (1, 1, 1), "adviser"),
    ]
    for key, label, basis, counts, unitname in units:
        ws5.append([label, basis])
        r = ws5.max_row
        for i, col in enumerate("IKM"):
            ws5.cell(row=r, column=3 + i).value = (
                f'=SUMIF(Budget!$Q${first}:$Q${last},"{key}",'
                f'Budget!${col}${first}:${col}${last})')
            ws5.cell(row=r, column=3 + i).number_format = MONEY
        for i in range(3):
            c = ws5.cell(row=r, column=6 + i)
            c.value = f"={get_column_letter(3+i)}{r}*{lf}"
            c.number_format = MONEY
        ws5.cell(row=r, column=9).value = f"=SUM(F{r}:H{r})"
        ws5.cell(row=r, column=9).number_format = MONEY
        ws5.cell(row=r, column=10).value = (
            f"=IF({counts[1]}=0,\"\",G{r}/{counts[1]})")
        ws5.cell(row=r, column=10).number_format = MONEY
        ws5.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws5.append([])
    ws5.append(["Counts used for the per-unit column", "Y1 3 teams / 4 focal points / 1 adviser; "
                "Y2-Y3 4 teams / 4 focal points / 1 adviser. Per-unit figures use year 2, "
                "the first full-strength year."])
    ws5.cell(row=ws5.max_row, column=2).alignment = Alignment(wrap_text=True)
    for col, w in zip("ABCDEFGHIJ", (38, 56, 13, 13, 13, 13, 13, 13, 14, 16)):
        ws5.column_dimensions[col].width = w

    wb.move_sheet("Summary", offset=-5)
    wb.save(path)
    return path


# --------------------------------------------------------------------------
# text version (computed in python from the same spec)
# --------------------------------------------------------------------------
def compute():
    rate = {r[0]: r[3] for r in RATES}
    unit = {r[0]: r[2] for r in RATES}
    a = {k: v for k, _l, v, _s in ASSUMPTIONS}
    rows = []
    for cat, act, item, basis, code, qty, h1, drv in LINES:
        tot = [q * rate[code] for q in qty]
        rows.append(dict(cat=cat, act=act, item=item, basis=basis, code=code,
                         unit=unit[code], rate=rate[code], qty=qty, tot=tot,
                         h1=h1, drv=drv))
    direct = [sum(r["tot"][i] for r in rows) for i in range(3)]
    goods = [sum(r["tot"][i] for r in rows if r["cat"] in GOODS_CATS) for i in range(3)]
    freight = [g * a["FREIGHT_PCT"] for g in goods]
    bank = [(direct[i] + freight[i]) * a["BANK_PCT"] for i in range(3)]
    sub = [direct[i] + freight[i] + bank[i] for i in range(3)]
    supp = [s * a["SUPPORT_PCT"] for s in sub]
    cont = [(sub[i] + supp[i]) * a["CONTING_PCT"] for i in range(3)]
    total = [sub[i] + supp[i] + cont[i] for i in range(3)]
    load = (total[0] / direct[0]) if direct[0] else 1.0
    # first six months (Oct 2026 - Mar 2027) as a share of year one
    h1_direct = sum(r["tot"][0] * r["h1"] for r in rows)
    h1_goods = sum(r["tot"][0] * r["h1"] for r in rows if r["cat"] in GOODS_CATS)
    h1_freight = h1_goods * a["FREIGHT_PCT"]
    h1_bank = (h1_direct + h1_freight) * a["BANK_PCT"]
    h1_sub = h1_direct + h1_freight + h1_bank
    h1_supp = h1_sub * a["SUPPORT_PCT"]
    h1_cont = (h1_sub + h1_supp) * a["CONTING_PCT"]
    h1_total = h1_sub + h1_supp + h1_cont
    # loaded cost of the delivery units
    drv = {}
    for key in ("TEAM", "FP", "HQ"):
        d = [sum(r["tot"][i] for r in rows if r["drv"] == key) for i in range(3)]
        drv[key] = dict(direct=d, loaded=[x * load for x in d])
    return rows, dict(direct=direct, freight=freight, bank=bank, sub=sub,
                      supp=supp, cont=cont, total=total, a=a, goods=goods,
                      load=load, h1=dict(direct=h1_direct, freight=h1_freight,
                                         bank=h1_bank, sub=h1_sub, supp=h1_supp,
                                         cont=h1_cont, total=h1_total),
                      drv=drv)


W = 80


def wrap(text, width=W, indent=""):
    import textwrap
    return textwrap.fill(text, width=width, initial_indent=indent,
                         subsequent_indent=indent)


def m(x):
    return f"{x:,.0f}"


def build_txt(path, xlsx_name):
    rows, t = compute()
    a = t["a"]
    o = []
    P = o.append
    P("=" * W)
    P("")
    P("        WHAT IT COSTS")
    P("        A three-year budget for the AP-RCA Priority Support Plan,")
    P("        Western South Sudan - lean, partner-delivered, seasonal")
    P("        Companion to PLAN_APRCA_WSS_ASSESSMENT_TO_ACTION_2026-08_EASY.txt")
    P(f"        August 2026 - INTERNAL. Live workbook: {xlsx_name}")
    P("")
    P("=" * W)
    P("")
    P("")
    P("THE BUDGET IN TEN LINES")
    P("-" * W)
    tot3 = sum(t["total"])
    P(f"  1  Three-year total      USD {m(tot3)}  ({m(t['total'][0])} / "
      f"{m(t['total'][1])} / {m(t['total'][2])})")
    P("  2  Shape                 No new organisation. A national partner")
    P("                           implements, SSWS leads visibly, Chinko HQ")
    P("                           oversees, AP South Sudan carries Juba.")
    P("  3  One expatriate        And only for 3 months a year, from Chinko -")
    P("                           not a resident post. 3% of the budget.")
    natl = sum(r["tot"][i] for r in rows for i in range(3)
               if r["cat"] in ("2 National staff", "3 Local partners and SSWS"))
    P(f"  4  Money on the ground   {natl/tot3:.0%} of the total is national staff,")
    P("                           partner grants and SSWS support.")
    P("  5  Seasonal payroll      Scouts are paid for the months they can work")
    P("                           (5 -> 7 -> 9), not for a calendar year.")
    P("  6  The corridor's price  Water, veterinary and meetings are budgeted")
    P("                           from Y2 - the herders named it, so it is a")
    P("                           line, not an intention.")
    P("  7  Y2 is the peak year   The aerial survey and the first borehole")
    P("                           both land there. Everything else is flat.")
    P("  8  Freight is 40%        Of everything imported. It is a line of its")
    P("                           own, not a rounding error.")
    P(f"  9  Support is a rate     {a['SUPPORT_PCT']:.0%} for Chinko HQ and the Juba desk -")
    P("                           services consumed, not a second structure.")
    P(" 10  Actions 2, 3 and 11   The three highest-value acts in the plan cost")
    P("                           nothing but staff time. They are already paid")
    P("                           for in the coordinator's salary.")
    P("")
    P("")
    P("THE OPERATION THIS BUDGET PAYS FOR")
    P("-" * W)
    for para in [
        "Four organisations already exist on or beside this ground, and the "
        "budget's first decision is to use them rather than duplicate them. A "
        "registered national NGO with offices in Wau and Raja implements: it "
        "holds the premises, hires the mobilisers, and is already inside the "
        "NGO Act's 80%-national staffing rule that would otherwise take a year "
        "to satisfy. The South Sudan Wildlife Service leads in public and on "
        "the ground; we equip its two existing but unequipped posts and pay "
        "allowances for joint missions, and we never pay a government salary. "
        "Chinko HQ provides technical oversight - three visits a year, not a "
        "resident expatriate. African Parks South Sudan carries Juba: banking, "
        "approvals, representation and the half-time finance officer sit in an "
        "office that already functions there.",
        "What that leaves us to pay for directly is small and deliberate: one "
        "national coordinator in Wau, one then two field liaison officers, "
        "seasonal scout teams, and the activities themselves. There is no "
        "vehicle fleet - 4x4s are hired by the day. There is no Juba office. "
        "There is no headquarters. The heaviest single asset in the budget is a "
        "motorbike.",
        "The calendar does the rest of the work. Every audience in this "
        "landscape is present November to February and absent June to "
        "September, so scouts are contracted seasonally - five months in year "
        "one, seven in year two, nine in year three - and the paper track runs "
        "in the rains from Juba, which is where the coordinator's and the "
        "lawyer's time goes when the field is shut. Budgeting twelve months of "
        "field payroll for a two-month window is the most common way to waste "
        "money here.",
    ]:
        P(wrap(para))
        P("")
    P("")
    P("THE HEADLINE")
    P("-" * W)
    P("")
    P(f"    {'':38s} {'YEAR 1':>12s} {'YEAR 2':>12s} {'YEAR 3':>12s} {'TOTAL':>12s}")
    P("    " + "-" * (W - 8))
    for cat in CATS:
        v = [sum(r["tot"][i] for r in rows if r["cat"] == cat) for i in range(3)]
        if not any(v):
            continue
        P(f"    {cat:38s} {m(v[0]):>12s} {m(v[1]):>12s} {m(v[2]):>12s} {m(sum(v)):>12s}")
    P("    " + "-" * (W - 8))
    P(f"    {'DIRECT COSTS':38s} {m(t['direct'][0]):>12s} {m(t['direct'][1]):>12s}"
      f" {m(t['direct'][2]):>12s} {m(sum(t['direct'])):>12s}")
    P(f"    {'Freight, customs, clearing (40%)':38s} {m(t['freight'][0]):>12s}"
      f" {m(t['freight'][1]):>12s} {m(t['freight'][2]):>12s} {m(sum(t['freight'])):>12s}")
    P(f"    {'Bank charges and FX (1.5%)':38s} {m(t['bank'][0]):>12s}"
      f" {m(t['bank'][1]):>12s} {m(t['bank'][2]):>12s} {m(sum(t['bank'])):>12s}")
    P(f"    {'Chinko HQ + Juba desk (8%)':38s} {m(t['supp'][0]):>12s}"
      f" {m(t['supp'][1]):>12s} {m(t['supp'][2]):>12s} {m(sum(t['supp'])):>12s}")
    P(f"    {'Contingency (7%)':38s} {m(t['cont'][0]):>12s}"
      f" {m(t['cont'][1]):>12s} {m(t['cont'][2]):>12s} {m(sum(t['cont'])):>12s}")
    P("    " + "=" * (W - 8))
    P(f"    {'TOTAL REQUESTED, USD':38s} {m(t['total'][0]):>12s} {m(t['total'][1]):>12s}"
      f" {m(t['total'][2]):>12s} {m(sum(t['total'])):>12s}")
    P("")
    P(wrap("Year 2 is the peak because two one-off items land in it: the aerial "
           "wildlife survey - the first since 2007, and the cheapest credibility "
           "the plan can buy - and the first corridor borehole. Strip those and "
           "the running cost of this operation is close to flat at around USD "
           f"{m((t['total'][0] + t['total'][2]) / 2)} a year."))
    P("")
    P("")
    P("THE SAME MONEY ON THE PLAN'S OWN CLOCK")
    P("-" * W)
    P(wrap("The assessment runs on 6 months, 1 year, 2 years, 3 years and on, "
           "because two clocks set every date in it: the FIELD clock has one "
           "window a year, 15 December to 15 February, and the PAPER clock runs "
           "all year from Juba. The budget is phased the same way. Note the "
           "first row is a slice of year one, not an addition to it."))
    P("")
    h = t["h1"]
    P(f"    {'PERIOD':38s} {'DIRECT':>12s} {'LOADED':>12s} {'CUMULATIVE':>12s}")
    P("    " + "-" * (W - 8))
    P(f"    {'6 MONTHS  Oct 2026 - Mar 2027':38s} {m(h['direct']):>12s}"
      f" {m(h['total']):>12s} {m(h['total']):>12s}")
    P(f"    {'1 YEAR    Oct 2026 - Sep 2027':38s} {m(t['direct'][0]):>12s}"
      f" {m(t['total'][0]):>12s} {m(t['total'][0]):>12s}")
    P(f"    {'2 YEARS   Oct 2027 - Sep 2028':38s} {m(t['direct'][1]):>12s}"
      f" {m(t['total'][1]):>12s} {m(t['total'][0]+t['total'][1]):>12s}")
    P(f"    {'3 YEARS   Oct 2028 onward':38s} {m(t['direct'][2]):>12s}"
      f" {m(t['total'][2]):>12s} {m(sum(t['total'])):>12s}")
    P("")
    P("  6 MONTHS - Oct 2026 to Mar 2027  |  GET LEGAL, GET SEEN, CHECK THE RIM")
    P(wrap(f"USD {m(h['total'])} loaded - {h['total']/t['total'][0]:.0%} of year "
           "one, and the only tranche that has to be committed before anyone "
           "knows whether season one deploys. It carries ten weeks of paper time "
           "in Juba (the coordinator, the legal retainer, registration, the "
           "regulation text and the s.24 gazette request), the partner grant's "
           "first half, the SSWS posts equipped, three scout teams recruited and "
           "trained, all the one-off kit, and season one itself: the rim "
           "reconnaissance flights, the field-checks, the first corridor talks. "
           "Almost every capital item sits here because a team without kit in "
           "December is a team that misses the year.", indent="    "))
    P("")
    P("  1 YEAR - to Sep 2027  |  TURN THE SEASON INTO INSTRUMENTS")
    P(wrap(f"USD {m(t['total'][0])} loaded for the full year. The second half is "
           "desk work and cheap: the wet season is when conservancy "
           "applications get written, the corridor package gets costed, and the "
           "survey gets designed and funded. The gate stays in the cash flow as "
           "well as in the plan: no SSWS backing by 30 November 2026 and the "
           "field half of this year should not be spent.", indent="    "))
    P("")
    P("  2 YEARS - Oct 2027 to Sep 2028  |  FIRST REAL PROTECTION")
    P(wrap(f"USD {m(t['total'][1])} loaded, the peak year, and it is peak for two "
           "reasons only: the aerial survey - the first since 2007 - and the "
           "first corridor borehole. Everything else is the same operation "
           "running longer: a fourth team, seven months instead of five, the "
           "veterinary campaign, the land-use workshops with the boom towns.",
           indent="    "))
    P("")
    P("  3 YEARS AND ON - Oct 2028 onward  |  DECLARE, THEN HOLD IT")
    P(wrap(f"USD {m(t['total'][2])} loaded, and this is the number that matters "
           "most for anyone thinking past the grant: it is roughly what holding "
           "this ground costs every year once the one-offs are done. Nine months "
           "of scouts, the partner grants, SSWS support, two reconnaissance "
           "flights, two boreholes and the paper. A declaration with no year-four "
           "money behind it is a map, not protection.", indent="    "))
    P("")
    P("")
    P("WHAT A TEAM AND A FOCAL POINT ACTUALLY COST - LOADED")
    P("-" * W)
    P(wrap("Direct cost is what the line items say. LOADED cost adds each unit's "
           "share of freight, bank charges, Chinko HQ and Juba support and "
           f"contingency - a factor of {t['load']:.2f} on year one. These are the "
           "numbers to quote when somebody asks what one ECHO/TANGO team costs "
           "to run, because they are what it costs to have one."))
    P("")
    d = t["drv"]
    P(f"    {'':38s} {'YEAR 1':>12s} {'YEAR 2':>12s} {'YEAR 3':>12s} {'3 YR':>12s}")
    P("    " + "-" * (W - 8))
    for key, label in (("TEAM", "ECHO/TANGO teams, all"),
                       ("FP", "Focal points + Wau/Juba backbone"),
                       ("HQ", "Chinko HQ oversight")):
        v = d[key]["loaded"]
        P(f"    {label:38s} {m(v[0]):>12s} {m(v[1]):>12s} {m(v[2]):>12s}"
          f" {m(sum(v)):>12s}")
    P("")
    tm = d["TEAM"]["loaded"]
    fp = d["FP"]["loaded"]
    P("    PER UNIT, AT FULL STRENGTH (year 2: four teams, four focal points)")
    P(f"      One ECHO/TANGO team, 4 scouts + 1 leader, 7 months     "
      f"{m(tm[1]/4):>10s}")
    P(f"      ...per scout-month on the ground                       "
      f"{m(tm[1]/4/5/7):>10s}")
    P(f"      One focal point (share of the Wau/Juba backbone)       "
      f"{m(fp[1]/4):>10s}")
    P(f"      Chinko HQ oversight, per year                          "
      f"{m(d['HQ']['loaded'][1]):>10s}")
    P("")
    P(wrap("Two things to read off that. First, a team is cheap and a season is "
           f"cheaper than a year: about USD {m(tm[1]/4)} buys five people, kitted, "
           "trained, mounted on a motorbike and present through the entire window "
           "in which anyone they need to reach is actually there. Second, the "
           "backbone costs more than the teams and should - it is the "
           "coordinator, the partner, the SSWS relationship and the Juba paper "
           "track, and it is the half of the operation that survives a cancelled "
           "season. A budget where the teams cost more than the backbone is a "
           "budget that has bought presence without the ability to convert it "
           "into an instrument."))
    P("")
    P(wrap("The ECHO and TANGO strands are not costed separately here, and "
           "deliberately: on this ground they are the same people doing awareness "
           "work and carrying the ground-truth checklist on the same patrol, "
           "under one registration as community scouts. Splitting them in the "
           "budget would create two payrolls and one team."))
    P("")
    P("")
    P("WHERE THE MONEY ACTUALLY GOES")
    P("-" * W)
    P("")
    shares = sorted(((sum(sum(r["tot"]) for r in rows if r["cat"] == c), c)
                     for c in CATS), reverse=True)
    for v, c in shares:
        if not v:
            continue
        bar = "#" * int(round(v / shares[0][0] * 34))
        P(f"    {c:34s} {v/sum(t['direct']):5.1%}  {bar}")
    P("")
    P(wrap("Read the top three together. People on the ground and the "
           "organisations that hold them - national staff, partner grants, SSWS "
           "support - are the operation. Field activities are the third block "
           "and they are dominated by two things that only fly: the survey and "
           "the reconnaissance flights over the mining cells. Everything else - "
           "premises, IT, transport, training - is under a tenth of the budget "
           "each, which is what it should look like when the office belongs to "
           "somebody else."))
    P("")
    P("")
    P("THE LINES, BY CATEGORY")
    P("-" * W)
    P(wrap("Quantities, not just totals: every number below is qty x unit cost. "
           "The workbook holds the same lines as live formulae."))
    for cat in CATS:
        sel = [r for r in rows if r["cat"] == cat]
        if not sel:
            continue
        P("")
        v = [sum(r["tot"][i] for r in sel) for i in range(3)]
        P(f"  {cat.upper()}   (3-yr {m(sum(v))})")
        P("  " + "-" * (W - 4))
        for r in sel:
            P(f"  {r['item']}")
            P(wrap(r["basis"], indent="      "))
            P(f"      {m(r['rate'])}/{r['unit']}   "
              f"Y1 {r['qty'][0]:g} = {m(r['tot'][0]):>9s}   "
              f"Y2 {r['qty'][1]:g} = {m(r['tot'][1]):>9s}   "
              f"Y3 {r['qty'][2]:g} = {m(r['tot'][2]):>9s}")
    P("")
    P("")
    P("WHAT EACH PLAN ACTION COSTS")
    P("-" * W)
    P("")
    for key in sorted(ACTIONS, key=lambda k: int(k[1:])):
        v = sum(sum(r["tot"]) for r in rows if r["act"] == key)
        P(f"    {ACTIONS[key]:44s} {m(v):>12s}")
    P("")
    P(wrap("Actions 3, 5 and 11 carry no line of their own, and that is the "
           "most important sentence in this budget. The s.24 gazette request "
           "that closes the ground to mineral-title applications, the finding "
           "that settles the road question, and the three funding conversations "
           "of week one are signatures, letters and phone calls. They are paid "
           "for inside the coordinator's salary and the legal retainer, and "
           "between them they are worth more than anything else here."))
    P("")
    P("")
    P("THE UNIT COSTS, AND HOW MUCH TO TRUST THEM")
    P("-" * W)
    P(wrap("Every rate is a planning assumption for this landscape, marked FIRM "
           "(expect to pay about this) or INDICATIVE (get a quote before you "
           "sign). None is a quotation and none is a promise. The workbook's "
           "Rates sheet is the only place they live: change one there and every "
           "line that uses it moves."))
    P("")
    for code, label, unit, usd, note in RATES:
        val = f"{usd:.0%}" if usd < 1 else m(usd)
        P(f"  {label}")
        P(f"      {val} per {unit}")
        P(wrap(note, indent="      "))
    P("")
    P("")
    P("THE FOUR RATES APPLIED TO EVERYTHING")
    P("-" * W)
    P("")
    for code, label, val, note in ASSUMPTIONS:
        P(f"  {label}: {val:.1%}")
        P(wrap(note, indent="      "))
        P("")
    P("")
    P("WHAT THIS BUDGET DELIBERATELY DOES NOT BUY")
    P("-" * W)
    for item, why in [
        ("A vehicle fleet",
         "Six months of the year the field is shut and a fleet still costs "
         "money, drivers and theft risk. 4x4s are hired by the day, for the "
         "days they are used."),
        ("An aircraft, or aircraft capability",
         "Two reconnaissance flights a year and one survey do not justify an "
         "airframe. Hours are bought, and the survey capability is borrowed "
         "from an operator who already flies this region to a published "
         "standard."),
        ("A Juba office",
         "There is a functioning one already, belonging to a partner in the "
         "same group. Renting a second one would buy nothing but a second "
         "lease."),
        ("A resident expatriate structure",
         "One adviser, three visits a year, from Chinko. Expatriate presence is "
         "the fastest way to turn a lean budget into a heavy one, and this "
         "landscape's bottleneck has never been technical knowledge."),
        ("Road rebuilding",
         "The assessment answered the road question: reopening the 1932 "
         "alignment means building ~180 km through uninhabited bush. That is a "
         "separate capital project with its own EIA, and it does not belong "
         "inside a governance budget."),
        ("Enforcement hardware",
         "No weapons, no vehicles for interception, no detention infrastructure. "
         "The plan's own posture is awareness first, seizure last; a budget that "
         "bought seizure capability in year one would contradict it."),
        ("Twelve-month scout contracts",
         "The audience is seasonal. Paying a full year for a two-month window is "
         "the most common way money disappears in this landscape."),
    ]:
        P(f"  * {item.upper()}")
        P(wrap(why, indent="    "))
    P("")
    P("")
    P("WHAT WOULD MAKE THIS BUDGET WRONG")
    P("-" * W)
    for item, why in [
        ("No SSWS backing by 30 November 2026",
         "Then season one does not deploy, and roughly two thirds of year one's "
         "field cost should not be spent. The gate is in the plan; keep it in "
         "the cash flow too."),
        ("A lost season",
         "One field window a year, in a state whose conflict record is quiet "
         "with bad days. A cancelled season pushes activity right by twelve "
         "months, it does not delete it - which is what the contingency is for "
         "and why the assets are movable."),
        ("A mineral title issued on the rim",
         "Then year two stops expanding and spends itself on that single "
         "problem. Cost is not the constraint there; sequencing is."),
        ("Partner capacity",
         "The whole model rests on a national partner being able to absorb and "
         "account for USD 60,000 a year. Verify that in the first Wau visit, "
         "before the grant is designed, not after."),
        ("Currency",
         "A meaningful share of this settles locally. The budget is in USD and "
         "carries a bank and FX line, but a sharp move in the local rate moves "
         "real purchasing power, not just the accounts."),
        ("Freight assumed away",
         "If a proposal reviewer cuts the 40% freight line to make the total "
         "look tidier, the equipment simply does not arrive. It is the line "
         "most often cut and least often survivable."),
    ]:
        P(f"  * {item.upper()}")
        P(wrap(why, indent="    "))
    P("")
    P("")
    P("HOW TO PHASE THE ASK")
    P("-" * W)
    P(wrap("This does not have to be raised as one number. It breaks cleanly "
           "into three, and the first is small enough to be somebody's "
           "discretionary grant:"))
    P("")
    paper = sum(sum(r["tot"]) for r in rows
                if r["cat"] == "11 Legal, registration and audit") / 3
    P(f"  A. THE PAPER PACKAGE - about USD {m(paper + 30000)} for year one")
    P(wrap("Coordinator, legal retainer, registration, and the Juba travel to "
           "file the regulation text and the gazette request. This buys the two "
           "actions with the highest value per dollar in the whole plan and "
           "needs no field season at all.", indent="     "))
    P("")
    P(f"  B. THE FIRST SEASON - about USD {m(t['total'][0])}")
    P(wrap("Everything in A plus three scout teams, the partner grant, SSWS "
           "support, the rim reconnaissance flights and the first corridor "
           "talks.", indent="     "))
    P("")
    P(f"  C. THE FULL THREE YEARS - USD {m(tot3)}")
    P(wrap("Adds the survey, the boreholes, the veterinary campaign, the fourth "
           "team and the land-use planning with the boom towns. This is the "
           "version that ends with a signed conservancy, a survey baseline and "
           "a fire number.", indent="     "))
    P("")
    P(wrap("On the funding side the assessment's action 11 still applies before "
           "any of this is circulated: ask whether any ground here is on the "
           "Keystone list of 162, take the seat at the EU conservation table in "
           "Juba rather than convening a rival one, and open the Southern "
           "National Park adjacency conversation with the incumbent operator "
           "next door. All three are free, and all three change what this "
           "budget should be attached to."))
    P("")
    P("")
    P("-" * W)
    P(wrap("HOW THIS FILE AND THE WORKBOOK RELATE. Both are generated from one "
           "specification, so they cannot disagree: the workbook is the live "
           "model (rates in one sheet, every cost a formula, subtotals as "
           "SUMIFs, support and contingency as percentage cells), and this file "
           "is the same model in prose. Edit rates and quantities in the "
           "workbook or the generator - never by typing over a total. Figures "
           "are USD. Quantities follow the assessment's field calendar (season "
           "one Dec 2026 - Feb 2027, season two Dec 2027 - Feb 2028, season "
           "three Dec 2028 - Feb 2029) and its eleven actions. Rates are "
           "planning assumptions for this landscape, not quotations: replace "
           "each INDICATIVE line with a real offer before signing anything."))
    P(f"Generated {datetime.date.today().isoformat()}.")
    P("=" * W)
    open(path, "w").write("\n".join(o) + "\n")
    return path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="reports/BUDGET_APRCA_WSS_2026-08_EASY.xlsx")
    ap.add_argument("--txt", default="reports/BUDGET_APRCA_WSS_2026-08_EASY.txt")
    a = ap.parse_args()
    build_xlsx(a.xlsx)
    import os
    build_txt(a.txt, os.path.basename(a.xlsx))
    rows, t = compute()
    print(f"wrote {a.xlsx} and {a.txt}")
    for i, y in enumerate(("Y1", "Y2", "Y3")):
        print(f"  {y}: direct {t['direct'][i]:>10,.0f}  total {t['total'][i]:>10,.0f}")
    print(f"  3-year total: {sum(t['total']):,.0f}")


if __name__ == "__main__":
    main()
